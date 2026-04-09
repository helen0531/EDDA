from fastapi import APIRouter, Request, Depends, Query, HTTPException, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from typing import Optional
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import sqlite3
import urllib.request
from urllib.parse import quote
import tempfile
import re
import os
import glob
import json
from PyPDF2 import PdfMerger
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors
from openpyxl.styles.colors import Color
from openpyxl.writer.excel import save_workbook
from sqlalchemy.orm import Session
from sqlalchemy import text

from ..db.database import get_db
from ..models.schemas import User
from ..services.auth_service import require_role, get_current_user

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

pdfmetrics.registerFont(TTFont('AppleGothic', '/System/Library/Fonts/Supplemental/AppleGothic.ttf'))

@router.get("/settings")
def settings_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager"], use_cache=False))
):
    settings_result = db.execute(text("SELECT value, start_date, end_date FROM settings WHERE key = 'max_overtime_hours'")).fetchone()
    max_overtime_hours = settings_result[0] if settings_result else 12
    start_date = settings_result[1] if settings_result else None
    end_date = settings_result[2] if settings_result else None
    return templates.TemplateResponse("settings.html", {"request": request, "current_user": current_user, "max_overtime_hours": max_overtime_hours, "start_date": start_date, "end_date": end_date})

@router.post("/settings")
def update_settings(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager"], use_cache=False)),
    max_overtime_hours: int = Form(...),
    start_date: str = Form(None),
    end_date: str = Form(None)
):
    db.execute(text("UPDATE settings SET value = :value, start_date = :start_date, end_date = :end_date WHERE key = 'max_overtime_hours'"), {"value": max_overtime_hours, "start_date": start_date, "end_date": end_date})
    db.commit()
    return templates.TemplateResponse("settings.html", {"request": request, "current_user": current_user, "max_overtime_hours": max_overtime_hours, "start_date": start_date, "end_date": end_date, "message": "설정이 저장되었습니다."})

@router.get("/stats")
def stats_page(
    request: Request,
    current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False))
):
    return templates.TemplateResponse("stats.html", {"request": request, "current_user": current_user})

@router.get("/api/stats/employee-status")
def get_employee_status(db: Session = Depends(get_db), current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False))):
    employees_result = db.execute(text("SELECT name FROM employees WHERE name != 'admin'")).fetchall()
    employees = [row[0] for row in employees_result]

    status_data = []
    for emp_name in employees:
        # Calculate remaining compensatory hours
        overtime_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '시간외 근무' AND status = 'approved'"), {"name": emp_name}).fetchall()
        total_overtime_hours = 0
        for req in overtime_requests:
            content = json.loads(req[0])
            total_overtime_hours += content.get('calculated_compensatory_hours', 0)

        leave_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '대휴 사용' AND status = 'approved'"), {"name": emp_name}).fetchall()
        used_leave_hours = 0
        for req in leave_requests:
            content = json.loads(req[0])
            used_leave_hours += content.get('hours', 0)

        remaining_hours = total_overtime_hours - used_leave_hours

        # Calculate remaining development cost
        dev_cost_requests = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '자기개발비' AND status = 'approved'"), {"name": emp_name}).fetchall()
        used_dev_cost = 0
        for req in dev_cost_requests:
            content = json.loads(req[0])
            used_dev_cost += int(content.get('cost', '0'))
        
        remaining_dev_cost = 2000000 - used_dev_cost

        status_data.append({
            "name": emp_name,
            "remaining_compensatory_hours": remaining_hours,
            "remaining_dev_cost": remaining_dev_cost
        })

    return JSONResponse(content=status_data)

@router.get("/api/stats/overtime-hours")
def get_overtime_hours(
    period: str = Query("monthly", enum=["monthly", "yearly"]),
    month: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False))
):
    employees_result = db.execute(text("SELECT name FROM employees WHERE name NOT IN ('admin', 'lead')")).fetchall()
    employees = [row[0] for row in employees_result]

    labels = []
    data = []
    today = datetime.today()

    for emp_name in employees:
        labels.append(emp_name)
        total_hours = 0

        if period == "monthly":
            if month:
                target_month_dt = datetime.strptime(month, "%Y-%m")
            else:
                target_month_dt = today
            
            end_date = (target_month_dt.replace(day=15)).strftime("%Y-%m-%d")
            start_date = (target_month_dt.replace(day=16) - relativedelta(months=1)).strftime("%Y-%m-%d")

        else: # yearly
            start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")
            end_date = today.replace(month=12, day=31).strftime("%Y-%m-%d")

        requests_result = db.execute(text("SELECT content FROM requests WHERE name = :name AND type = '시간외 근무' AND status = 'approved' AND created BETWEEN :start_date AND :end_date"), {"name": emp_name, "start_date": start_date, "end_date": end_date}).fetchall()
        for req in requests_result:
            content = json.loads(req[0])
            total_hours += content.get('work_hours_weekday', 0) + content.get('work_hours_holiday', 0)
        
        data.append(total_hours)

    return JSONResponse(content={"labels": labels, "data": data})

@router.get("/admin-dashboard")
def admin_dashboard(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_name: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False)),
    overtime_page: int = 1,
    compensatory_page: int = 1,
    trip_page: int = 1,
    dev_page: int = 1,
    per_page: int = 10
):
    # Fetch unique names for the dropdown
    names_result = db.execute(text("SELECT DISTINCT name FROM requests ORDER BY name ASC")).fetchall()
    names = [row[0] for row in names_result]

    # Date filtering logic
    if not start_date or not end_date:
        today = datetime.today()
        if today.day < 16:
            start_of_month = today.replace(day=16) - relativedelta(months=1)
        else:
            start_of_month = today.replace(day=16)
        end_of_month = start_of_month + relativedelta(months=1) - relativedelta(days=1)
        
        start_date = start_of_month.strftime("%Y-%m-%d")
        end_date = end_of_month.strftime("%Y-%m-%d")

    # Base query and params
    base_query = "SELECT * FROM requests WHERE created BETWEEN :start_date AND :end_date"
    params = {"start_date": start_date, "end_date": end_date}

    if selected_name and selected_name != "all":
        base_query += " AND name = :name"
        params["name"] = selected_name

    # Pagination
    overtime_offset = (overtime_page - 1) * per_page
    compensatory_offset = (compensatory_page - 1) * per_page
    trip_offset = (trip_page - 1) * per_page
    dev_offset = (dev_page - 1) * per_page

    # Initialize request lists
    overtime_requests = []
    overtime_allowance_requests = []
    overtime_compensatory_requests = []
    trip_requests = []
    self_dev_rows = []
    compensatory_leave_requests = []
    total_compensatory = 0
    total_trip = 0
    total_dev = 0
    total_overtime = 0
    total_overtime_allowance = 0
    total_overtime_compensatory = 0

    if not request_type or request_type == "all" or request_type.startswith("시간외 근무"):
        overtime_query = f"{base_query} AND type = '시간외 근무'"
        if request_type == "시간외 근무 - 수당지급":
            overtime_query += " AND json_extract(content, '$.compensation') = '수당지급'"
        elif request_type == "시간외 근무 - 대체휴가":
            overtime_query += " AND json_extract(content, '$.compensation') = '대체휴가'"
        overtime_query += " ORDER BY id DESC LIMIT :limit OFFSET :offset"
        overtime_requests_raw = db.execute(text(overtime_query), {**params, "limit": per_page, "offset": overtime_offset}).fetchall()
        
        overtime_requests = []
        for r in overtime_requests_raw:
            content = json.loads(r[3])
            r = list(r)
            r.append(content.get('compensation'))
            overtime_requests.append(r)

        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '시간외 근무'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"

        if request_type == "시간외 근무 - 수당지급":
            count_query += " AND json_extract(content, '$.compensation') = '수당지급'"
        elif request_type == "시간외 근무 - 대체휴가":
            count_query += " AND json_extract(content, '$.compensation') = '대체휴가'"
        total_overtime = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type in ["대휴사용", "대휴신청"]:
        compensatory_leave_query = f"{base_query} AND type IN ('대휴 사용', '대휴신청') ORDER BY id DESC LIMIT :limit OFFSET :offset"
        compensatory_leave_requests = db.execute(text(compensatory_leave_query), {**params, "limit": per_page, "offset": compensatory_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type IN ('대휴 사용', '대휴신청')"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_compensatory = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type == "출장":
        trip_query = f"{base_query} AND type = '출장' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        trip_requests = db.execute(text(trip_query), {**params, "limit": per_page, "offset": trip_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '출장'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_trip = db.execute(text(count_query), params).scalar_one()

    if not request_type or request_type == "all" or request_type == "자기개발비":
        dev_query = f"{base_query} AND type = '자기개발비' ORDER BY id DESC LIMIT :limit OFFSET :offset"
        self_dev_rows = db.execute(text(dev_query), {**params, "limit": per_page, "offset": dev_offset}).fetchall()
        count_query = f"SELECT COUNT(*) FROM requests WHERE created BETWEEN :start_date AND :end_date AND type = '자기개발비'"
        if selected_name and selected_name != "all":
            count_query += " AND name = :name"
        total_dev = db.execute(text(count_query), params).scalar_one()

    overtime_total_pages = (total_overtime + per_page - 1) // per_page
    compensatory_total_pages = (total_compensatory + per_page - 1) // per_page
    trip_total_pages = (total_trip + per_page - 1) // per_page
    dev_total_pages = (total_dev + per_page - 1) // per_page

    pending_count = db.execute(text("SELECT count(*) FROM requests WHERE status LIKE '%대기'")).scalar_one()
    approved_count = db.execute(text("SELECT count(*) FROM requests WHERE status = 'approved'")).scalar_one()
    rejected_count = db.execute(text("SELECT count(*) FROM requests WHERE status = 'rejected'")).scalar_one()
    total_count = db.execute(text("SELECT count(*) FROM requests")).scalar_one()
    progress_count = total_count - (pending_count + approved_count + rejected_count)

    total_items = 0
    if not request_type or request_type == "all":
        total_items = total_overtime + total_compensatory + total_trip + total_dev
    elif request_type.startswith("시간외 근무"):
        total_items = total_overtime
    elif request_type in ["대휴사용", "대휴신청"]:
        total_items = total_compensatory
    elif request_type == "출장":
        total_items = total_trip
    elif request_type == "자기개발비":
        total_items = total_dev

    total_pages = (total_items + per_page - 1) // per_page



    return templates.TemplateResponse(
        "admin_dashboard.html",
        {
            "request": request,
            "overtime_requests": overtime_requests,
            "compensatory_leave_requests": compensatory_leave_requests,
            "trip_requests": trip_requests,
            "dev_requests": self_dev_rows,
            "pending_count": pending_count,
            "progress_count": progress_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "current_user": current_user,
            "start_date": start_date,
            "end_date": end_date,
            "names": names,
            "selected_name": selected_name,
            "request_type": request_type,
            "now": datetime.now(),
            "overtime_current_page": overtime_page,
            "overtime_total_pages": overtime_total_pages,
            "compensatory_current_page": compensatory_page,
            "compensatory_total_pages": compensatory_total_pages,
            "trip_current_page": trip_page,
            "trip_total_pages": trip_total_pages,
            "dev_current_page": dev_page,
            "dev_total_pages": dev_total_pages
        }
    )

@router.get("/admin-dashboard/pdf/merge")
def merge_pdfs(ids: str, db: Session = Depends(get_db), current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False))):
    request_ids = [int(id) for id in ids.split(',')]

    request_ids_str = ",".join(map(str, request_ids))
    query = f"""SELECT r.id 
                 FROM requests r JOIN employees e ON r.name = e.name 
                 WHERE r.id IN ({request_ids_str}) 
                 ORDER BY e.emp_no ASC, json_extract(r.content, '$.work_date') ASC"""
    sorted_requests_result = db.execute(text(query)).fetchall()
    sorted_request_ids = [row[0] for row in sorted_requests_result]

    merger = PdfMerger()
    
    pdf_contents = []
    has_compensatory_leave = False
    for request_id in sorted_request_ids:
        request_data_result = db.execute(text("SELECT type, content FROM requests WHERE id = :id"), {"id": request_id}).fetchone()
        if request_data_result:
            request_type = request_data_result[0]
            request_content = json.loads(request_data_result[1])
            if (request_type == '시간외 근무' and request_content.get('compensation') == '대체휴가') or request_type in ['대휴신청', '대휴 사용']:
                has_compensatory_leave = True
                continue
        
        pdf_content = download_pdf(request_id, db, current_user)
        pdf_contents.append(BytesIO(pdf_content))

    if has_compensatory_leave:
        return HTMLResponse(content="<script>alert('대휴 신청은 별도 문서를 생성하지 않습니다.'); window.history.back();</script>")

    if not pdf_contents:
        return HTMLResponse(content="선택된 PDF가 없습니다.", status_code=404)

    for pdf_content in pdf_contents:
        merger.append(pdf_content)

    output_buffer = BytesIO()
    merger.write(output_buffer)
    merger.close()

    output_buffer.seek(0)
    return HTMLResponse(content=output_buffer.read(), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=merged-report.pdf"})

@router.get("/admin-dashboard/excel")
def export_to_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    selected_name: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager", "lead"], use_cache=False))
):
    if request_type:
        request_type = request_type.strip()
    
    if request_type != '시간외 근무 - 수당지급':
        return HTMLResponse(content="<script>alert('시간외 근무 - 수당지급을 선택하고 엑셀로 내보내기를 해주세요.'); window.history.back();</script>")

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "시간외 및 휴일근무 내역"
    
    # Get the current year and month for the title
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        title_date = start_date_obj.strftime('%Y년 %m월')
    else:
        title_date = datetime.now().strftime('%Y년 %m월')
    
    # Add title row with merged cells
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=f"{title_date} 시간외 및 휴일근무 내역")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Malgun Gothic', size=18, bold=True, color='FFFF0000')
    
    # Set row height for title row
    ws.row_dimensions[1].height = 36.75
    
    # Add empty row
    ws.append([])
    
    # Set headers at row 3
    headers = ["번호", "성명", "직번", "직급", "시간외 근로유형", "근무일", "요일", "근무시간 from", "근무시간 to", "부서장 인정시간", "연장근무 세부내역"]
    ws.append(headers)
    
    # Add another row for headers and merge cells except H and I
    subheaders = ["", "", "", "", "", "", "", "시작", "종료", "", ""]
    ws.append(subheaders)
    
    # Merge cells for headers (row 3 and 4) except columns H and I
    for col in [1, 2, 3, 4, 5, 6, 7, 10, 11]:  # A, B, C, D, E, F, G, J, K
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
    
    # Set specific text for H3 and I3
    ws.cell(row=3, column=8).value = "근무시간"
    ws.cell(row=3, column=9).value = "근무시간"
    
    # Merge H3 and I3 horizontally
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)
    
    # Set header style (rows 3 and 4): white bold text on dark blue background
    # In openpyxl, colors are specified in RRGGBB format (without # prefix)
    # For white text, we use FFFFFF
    # header_font = Font(name='Malgun Gothic', size=20, bold=True, color='FFFFFFFF')
    
    # 헤더 스타일 정의
    header_fill = PatternFill(start_color='FFffffff', end_color='FFffffff', fill_type='solid')
    # header_fill = PatternFill(start_color='FF081F5C', end_color='FF081F5C', fill_type='solid')
    header_font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFFFF')
    
    # 헤더 셀에 스타일 적용
    for row in range(3, 5):  # Rows 3 and 4
        for col in range(1, 12):  # Columns A through K
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font

    # Fetch data
    where_clauses = []
    params = {}

    if start_date and end_date:
        where_clauses.append("r.created BETWEEN :start_date AND :end_date")
        params["start_date"] = start_date
        params["end_date"] = end_date

    where_clauses.append("r.status = 'approved'")
    where_clauses.append("r.type = '시간외 근무'")
    where_clauses.append("json_extract(r.content, '$.compensation') = '수당지급'")

    if selected_name and selected_name != "all":
        where_clauses.append("r.name = :name")
        params["name"] = selected_name

    where_clause = " AND ".join(where_clauses)
    query = f"""SELECT r.*, e.emp_no, e.position 
                 FROM requests r JOIN employees e ON r.name = e.name 
                 WHERE {where_clause} ORDER BY e.emp_no ASC, json_extract(r.content, '$.work_date') ASC"""

    requests_result = db.execute(text(query), params).fetchall()

    # Group requests by name
    from itertools import groupby
    def get_name(row):
        return row._mapping['name']

    grouped_requests = {k: list(v) for k, v in groupby(sorted(requests_result, key=get_name), key=get_name)}

    row_num = 5  # Start data from row 5 (after title row, empty row, header row, and subheader row)
    unique_id = 1
    for name, requests in grouped_requests.items():
        total_hours = 0
        start_row = row_num  # Remember the starting row for this name group
        
        for i, row_mapping in enumerate(requests):
            row = row_mapping._mapping
            content = json.loads(row['content'])
            work_date_str = content.get('work_date')
            
            day_of_week_korean = ''
            formatted_date = ''
            display_work_type = ''
            if work_date_str:
                try:
                    work_date = datetime.strptime(work_date_str, '%Y-%m-%d')
                    formatted_date = work_date.strftime('%Y.%m.%d')
                    day_map = {"Mon": "월", "Tue": "화", "Wed": "수", "Thu": "목", "Fri": "금", "Sat": "토", "Sun": "일"}
                    day_of_week_korean = day_map.get(work_date.strftime('%a'), '')

                    # 근무 유형을 실제 요일 기준으로 판단
                    if work_date.weekday() < 5: # 0:월, 1:화, 2:수, 3:목, 4:금
                        display_work_type = "평일연장근로(최대3시간)"
                    else: # 5:토, 6:일
                        display_work_type = "휴일근로(최대6시간)"
                except ValueError:
                    pass

            work_time_range = content.get('work_time_range', '-').split('-')
            
            if i == 0:
                ws.cell(row=row_num, column=1, value=unique_id)
                unique_id += 1
            else:
                ws.cell(row=row_num, column=1, value="")

            ws.cell(row=row_num, column=2, value=row['name'])
            ws.cell(row=row_num, column=3, value=row['emp_no'])
            ws.cell(row=row_num, column=4, value=row['position'])
            ws.cell(row=row_num, column=5, value=display_work_type)
            ws.cell(row=row_num, column=6, value=formatted_date)
            ws.cell(row=row_num, column=7, value=day_of_week_korean)
            ws.cell(row=row_num, column=8, value=work_time_range[0])
            ws.cell(row=row_num, column=9, value=work_time_range[1] if len(work_time_range) > 1 else '')
            hours = content.get('work_hours_weekday', 0) + content.get('work_hours_holiday', 0)
            ws.cell(row=row_num, column=10, value=hours)
            total_hours += hours
            ws.cell(row=row_num, column=11, value=content.get('reason_detail', ''))
            row_num += 1

        # Merge cells in column B (name) if there are multiple rows for this name
        if row_num > start_row + 1:  # If there's more than one row for this name
            ws.merge_cells(start_row=start_row, start_column=2, end_row=row_num-1, end_column=2)
        
        # Add summary row
        ws.cell(row=row_num, column=10, value=total_hours)
        row_num += 1

    # Set font and alignment for all cells
    default_font = Font(name='Malgun Gothic', size=9)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(ws.iter_rows(), 1):  # row_idx starts from 1
        for cell in row:
            # Skip title row (row 1) to preserve its font settings
            if row_idx != 1:
                cell.font = default_font
            if cell.column != 11:  # "연장근무 세부내역" column
                cell.alignment = center_alignment

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "시간외_및_휴일근무_내역.xlsx"

    return HTMLResponse(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('시간외_및_휴일근무_내역.xlsx')}"}
    )

def download_pdf(request_id: int, db: Session, current_user: User) -> bytes:
    request_data_result = db.execute(text("SELECT * FROM requests WHERE id = :id"), {"id": request_id}).fetchone()

    if not request_data_result:
        return HTMLResponse(content="Request not found", status_code=404)
    
    request_data = dict(request_data_result._mapping)

    signature_result = db.execute(text("SELECT signature FROM employees WHERE name = :name"), {"name": '최혜영'}).fetchone()
    signature = signature_result[0] if signature_result else None

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    styles['h1'].fontName = 'AppleGothic'
    styles['h1'].alignment = TA_CENTER
    styles['h2'].fontName = 'AppleGothic'
    styles['Normal'].fontName = 'AppleGothic'
    styles['Normal'].fontSize = 12

    # Title
    title = request_data['type']
    if title == '시간외 근무':
        title = '시간외 근무'
    story.append(Paragraph(f"{title} 승인 확인서", styles['h1']))
    story.append(Spacer(1, 24))

    # Request Details
    request_data_list = [[Paragraph(f"<b>신청자:</b> {request_data['name']}", styles['Normal'])], [Paragraph(f"<b>신청일:</b> {request_data['created'].split(' ')[0]}", styles['Normal'])]]
    request_content = json.loads(request_data['content'])
    if request_data['type'] == '시간외 근무' and request_content.get('compensation') == '대체휴가':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        doc.build([Paragraph("대체휴가 신청은 별도의 승인 확인서가 발급되지 않습니다.", styles['Normal'])])
        buffer.seek(0)
        return buffer.read()
    key_map = {
        "work_type": "근무 유형", "work_date": "근무일", "work_hours_weekday": "평일 근무시간",
        "work_hours_holiday": "휴일 근무시간", "reason_type": "신청 사유", "reason_detail": "상세 사유",
        "work_location": "근무 장소", "compensation": "보상 유형", "course_title": "수강 항목",
        "purpose": "목적", "course_content": "수강 내용", "cost": "비용", "start_date": "시작일",
        "end_date": "종료일", "reference_site": "참고 사이트", "region": "출장 지역",
        "organization": "출장 기관", "transport": "이동 수단", "hours": "사용 시간"
    }
    for key, value in request_content.items():
        if key in key_map:
            request_data_list.append([Paragraph(f"<b>{key_map[key]}:</b> {value}", styles['Normal'])])

    request_table = Table(request_data_list, hAlign='LEFT')
    request_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'AppleGothic'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    story.append(request_table)

    story.append(Spacer(1, 48))

    approval_text = "상기 신청을 승인합니다."
    if request_data['type'] == '시간외근무':
        approval_text = "상기 시간외 근무 신청을 승인합니다."
    elif request_data['type'] == '출장':
        approval_text = "상기 출장 신청을 승인합니다."
    elif request_data['type'] == '자기개발비':
        approval_text = "상기 자기개발비 신청을 승인합니다."
    story.append(Paragraph(approval_text, styles['Normal']))
    story.append(Spacer(1, 24))

    # Approval Info
    story.append(Paragraph(f"<b>승인 날짜:</b> {request_data['created'].split(' ')[0]}", styles['Normal']))
    story.append(Spacer(1, 12)) # Add space
    story.append(Paragraph(f"<b>승인자:</b> 팀장 최혜영", styles['Normal']))
    if signature and signature['signature']:
        story.append(Image(signature['signature'], width=50, height=50, hAlign='LEFT'))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()

@router.get("/admin-dashboard/pdf/{request_id}")
def download_pdf_route(request_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    request_owner_result = db.execute(text("SELECT name FROM requests WHERE id = :id"), {"id": request_id}).fetchone()
    request_owner = request_owner_result[0] if request_owner_result else None

    if not request_owner or (current_user.name != request_owner and current_user.role not in ["admin", "manager", "lead"]):
        raise HTTPException(status_code=403, detail="이 PDF에 접근할 권한이 없습니다.")

    pdf_content = download_pdf(request_id, db, current_user)
    return HTMLResponse(content=pdf_content, media_type="application/pdf")
